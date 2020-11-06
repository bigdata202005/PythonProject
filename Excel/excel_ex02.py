from openpyxl import Workbook
# Workbook 오브젝트 생성
wb = Workbook()
print(wb)
# 워크시트 생성하기
ws1 = wb.create_sheet("Mysheet1")       # insert at the end (default)
ws2 = wb.create_sheet("Mysheet2", 0)    # insert at first position
ws3 = wb.create_sheet("Mysheet3", -1)   # insert at the penultimate position
print(ws1, ws2, ws3)
# 저장하기 : 빈문서(시트 4개)
wb.save("ex02.xlsx")
