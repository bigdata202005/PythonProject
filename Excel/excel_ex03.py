from openpyxl import Workbook
# Workbook 오브젝트 생성
wb = Workbook()
print(wb)

# 워크시트 생성하기
ws1 = wb.create_sheet("임시작업")       # insert at the end (default)
ws1.sheet_properties.tabColor = "1072BA" # 워크시트 탭 색상 지정

# 활성 시트 선택
ws = wb.active
# 탭색상변경
ws.sheet_properties.tabColor = "FF0000"
# 시트이름변경
ws.title = "작업시트"

# 워크시트 선택
ws = wb["임시작업"]
# 속성 표시
print(ws.sheet_properties)
# 시트 이름 보기
print(ws.title)

# 시트이름 전체보기
print(wb.sheetnames)
# 시트별로 작업
for w in wb:
    print(w.title)  # 이름보기

# 워크시트 복사하기
source = wb.active
wb.copy_worksheet(source)

# 시트이름 전체보기
print(wb.sheetnames)

# 저장하기 : 빈문서(시트 3개)
wb.save("ex03.xlsx")
