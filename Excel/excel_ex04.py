from openpyxl import Workbook

# Workbook 오브젝트 생성
wb = Workbook()

# 시트이름 전체보기
print(wb.sheetnames)
ws = wb.active
ws.title = '셀작업'
print(wb.sheetnames)
# 셀에 값쓰기
c = ws['A3']
c.value = 123
ws['b3'] = 111
# 4행 2열에 값 10 할당
ws.cell(row=4, column=2, value=10)

# 반복문으로 데이터 작성
ws = wb.create_sheet("반복작업")
ws = wb["반복작업"]
print(ws.title)
# 10행 10열에 데이터 넣기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=x * y)

# 가로합/세로합 구하기
# 셀값 초기화
for x in range(1,11):
    ws.cell(row=x, column=11, value=0)
    ws.cell(row=11, column=x, value=0)
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=11).value += x * y
        ws.cell(row=11, column=y).value += x * y


# 저장하기 : 빈문서(시트 1개)
wb.save("ex04.xlsx")
