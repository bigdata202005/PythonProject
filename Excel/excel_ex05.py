from openpyxl import Workbook

# Workbook 오브젝트 생성
wb = Workbook()

ws = wb.active
ws.title = '셀범위 작업'
print(wb.sheetnames)
# cell 범위 지정
cell_range = ws['A1':'C2']
print(cell_range)
print("~" * 80)
i = 0
for row in cell_range:
    for cell1 in row:
        i += 1
        cell1.value = i
        print(cell1, cell1.value)
print("~" * 80)


ws1 = wb.create_sheet("열행선택")
# C1
colC = ws1['C']
print("ws1['C'] : ", colC)
colC[0].value = 1234
print("~" * 80)
# A10:C10
row10 = ws1[10]
print("ws1[10] : ", row10)
for cell2 in row10:
    i += 1
    cell2.value = i
print("~" * 80)
# D1:E10
col_range = ws1['D:E']
print("ws1['D:E'] : ", col_range)
for cell3 in col_range:
    i += 1
    cell3[0].value = i
print("~" * 80)
# A13:E15
row_range = ws1[13:15]
print("ws1[13:15] : ", row_range)
for row in row_range:
    for cell4 in row:
        i += 1
        cell4.value = i
        print(cell4, cell4.value)
print("~" * 80)

# 저장하기
wb.save("ex05.xlsx")
