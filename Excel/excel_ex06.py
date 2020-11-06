from openpyxl import Workbook

# Workbook 오브젝트 생성
wb = Workbook()

ws = wb.active
ws.title = '워크시트1'
print(wb.sheetnames)
i = 0
# ws.iter_rows
# A1:C2
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        i += 2
        cell.value = i
        print(cell, cell.value)
print("~" * 80)
# 성능상의 이유로이 Worksheet.iter_cols()메서드는 읽기 전용 모드에서 사용할 수 없습니다.
# A3:C5
for col in ws.iter_cols(min_row=3, max_col=3, max_row=5):
    for cell in col:
        i += 2
        cell.value = i
        print(cell, cell.value)
print("~" * 80)

for row in ws.iter_rows(min_row=1, min_col=2, max_col=3, max_row=2):
    for cell in row:
        i += 2
        cell.value = i
        print(cell, cell.value)
print("~" * 80)
# 저장하기
wb.save("ex06.xlsx")
