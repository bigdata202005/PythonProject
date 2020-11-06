from openpyxl import Workbook

# Workbook 오브젝트 생성
wb = Workbook()

ws = wb.active
ws.title = '워크시트1'
print(wb.sheetnames)
ws['C9'] = 'hello world'
# 파일의 모든 행 또는 열을 반복해야하는 경우 대신 다음 Worksheet.rows속성을 사용할 수 있습니다.
print(tuple(ws.rows))
print("~" * 80)
for row in tuple(ws.rows):
    for cell in row:
         print(cell, cell.value)
print("~" * 80)
# 또는 Worksheet.columns속성 :
# 성능상의 이유로 Worksheet.columns읽기 전용 모드 에서는 속성을 사용할 수 없습니다.
print(tuple(ws.columns))
print("~" * 80)
for columns in tuple(ws.columns):
    for cell in columns:
         print(cell, cell.value)
print("~" * 80)
# 저장하기
wb.save("ex07.xlsx")
