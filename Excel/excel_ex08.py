from openpyxl import Workbook
import random

# Workbook 오브젝트 생성
wb = Workbook()

ws = wb.active
ws.title = '워크시트1'
print(wb.sheetnames)

# 10행 10열에 데이터 넣기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=random.randint(0, 100))
print("~" * 80)

# 값만
# 워크 시트의 값만 원하는 경우 Worksheet.values속성을 사용할 수 있습니다 .
# 이렇게하면 워크 시트의 모든 행이 반복되지만 셀 값만 반환됩니다.
for row in ws.values:
    for value in row:
        print("{:4d}".format(value), end=' ')
    print()
print("~" * 80)

# Worksheet.iter_rows()및 둘 다 매개 변수를 사용하여 셀의 값만 반환 Worksheet.iter_cols()할 수 있습니다
# values_only.
for row in ws.iter_rows(min_row=1, max_col=10, max_row=10, values_only=True):
    print(row)
    for value in row:
        print("{:4d}".format(value), end=' ')
    print()
print("~" * 80)
# 저장하기
wb.save("ex08.xlsx")
