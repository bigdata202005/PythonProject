from openpyxl import Workbook
import random
from tempfile import NamedTemporaryFile

# Workbook 오브젝트 생성
wb = Workbook()

ws = wb.active
ws.title = '워크시트1'
print(wb.sheetnames)

# 10행 10열에 데이터 넣기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=random.randint(0, 100))
print("~" * 80)

# 저장하기
# wb.save("ex09_1.xlsx")
# 예를 들어 Pyramid, FlaskExam 또는 Django와 같은 웹 애플리케이션을 사용할 때
# 파일을 스트림에 저장하려면 다음을 제공하면됩니다 NamedTemporaryFile()
tmp = None
try:
    tmp = NamedTemporaryFile()
    file_name = tmp.name
    print(file_name)
    wb.save(file_name)
    tmp.seek(0)
    stream = tmp.read()
    print(stream)
except PermissionError as e:
    print(e)
finally:
    tmp.close()
    print("Bye!!!")

