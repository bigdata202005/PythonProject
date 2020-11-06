from openpyxl import load_workbook
# 엑셀 읽기
wb = load_workbook('ex08.xlsx')
print(wb.sheetnames)
print("~" * 80)
# 모든 데이터 읽기
ws = wb.active
for row in ws.values:
    for value in row:
        print("{:4d}".format(value), end=' ')
    print()
print("~" * 80)
