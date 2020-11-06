import json

from openpyxl import load_workbook
import xlrd
# 엑셀 읽기
wb = load_workbook('lotto2.xlsx')
# wb = xlrd.open_workbook('lotto.xls')
print(wb.sheetnames)
ws = wb.active

# B, N, o,p,q,r,s,t
# 4~932
lotto_dict = dict()
for row in range(4, 933):
    key = ws['B' + str(row)].value
    lotto_list = []
    for col in ['N','O','P','Q','R','S','T']:
        lotto_list.append(ws[col + str(row)].value)
    lotto_dict[key] = lotto_list
print(lotto_dict)
with open("lotto.json", "w") as json_file:
    json.dump(lotto_dict, json_file)
    print("저장 완료!!!")