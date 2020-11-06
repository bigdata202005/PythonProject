# Excel 파일 다루기
# pip install openpyxl  ==> 설치
from openpyxl import Workbook
# 엑셀문서 만들기
wb = Workbook()

# 워크시트 추가하기
ws1 = wb.create_sheet("Mysheet1")       # 뒤에 추가
ws2 = wb.create_sheet("Mysheet2", 0)    # 0번째 추가
ws3 = wb.create_sheet("Mysheet3", -1)   # 뒤에서 두번째 추가

# 저장하기
wb.save("Ex02.xlsx")
