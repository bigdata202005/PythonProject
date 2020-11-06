# Excel 파일 다루기
# pip install openpyxl  ==> 설치
from openpyxl import Workbook
# 엑셀문서 만들기
wb = Workbook()

# 워크시트 추가하기
ws1 = wb.create_sheet("Mysheet1")       # 뒤에 추가
ws2 = wb.create_sheet("Mysheet2", 0)    # 0번째 추가
ws3 = wb.create_sheet("Mysheet3", -1)   # 뒤에서 두번째 추가

# 시트이름 변경하기
ws1.title = "변경된 이름"
# 시트의 탭 색상 변경
ws1.sheet_properties.tabColor = "1072BA"
ws2.sheet_properties.tabColor = "FF0000"
# 모든 시트이름 보기
print(wb.sheetnames)

# 현재 시트 선택
ws = wb.active
# 시트이름
print(ws.title)
# 시트 변경
ws = wb["Mysheet3"]
# 시트이름
print(ws.title)

# 저장하기
wb.save("Ex03.xlsx")
