# Excel 파일 다루기
# pip install openpyxl  ==> 설치
from openpyxl import Workbook

# 엑셀문서 만들기
wb = Workbook()
print(wb.sheetnames)
ws = wb.active
ws.title = "이름변경"
print(wb.sheetnames)

# 시트 복사
target = wb.copy_worksheet(ws)
print(wb.sheetnames)
target.title = "복사된시트"
print(wb.sheetnames)

# 데이터 추가!!!!
cell = ws['A4']
cell.value = 123
print(ws['A4'].value)
print(cell.value)

ws['A5'] = 100
print(ws['A5'].value)

ws.cell(row=4, column=2, value=10)
print(ws['B4'].value)
# 1줄 입력
target.append(['이름', '나이', '주소', '성별'])
target.append(['한사람', 22, '서울', '남자'])

# 반복문으로 추가
for r in range(3, 10):
    for c in range(1, 4):
        target.cell(row=r, column=c, value=r*c)

# 저장하기
wb.save("Ex04.xlsx")
