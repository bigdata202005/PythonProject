# Excel 파일 다루기
# pip install openpyxl  ==> 설치
from openpyxl import Workbook

# 엑셀문서 만들기
wb = Workbook()
print(wb.sheetnames)
ws = wb.active

# 범위지정
cell_range = ws['A1':'C2']
print("ws['A1':'C2'] :", cell_range)
print('~' * 80)
i = 0
r = 0
for row in cell_range:
    sum = 0  # 가로 합
    r += 1
    for cell in row:
        i += 1
        cell.value = i
        sum += cell.value
        print(cell, end=' ')
    ws['d'+str(r)] = sum
    print()
print('~' * 80)

cell_range = ws['A1':'D2']
for row in cell_range:
    for cell in row:
       print(cell.value, end=' ')
    print()
print('~' * 80)

# 행은 이전에 작업했던 행까지만 적용
colC = ws['C']  # 열선택 : ws['c1':'c2']
print(colC)
col_range = ws['C:D'] # ws['c1':'d2']
print(col_range)

# 열은 이전에 작업했던 열까지만 적용
row10 = ws[10] # 행선택 : ws['a10':'d10']
print(row10)
row_range = ws[15:20] # 행선택 : ws['a15':'d20']
print(row_range)

# 저장하기
wb.save("Ex05.xlsx")
