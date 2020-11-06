# Excel 파일 다루기
# pip install openpyxl  ==> 설치
from openpyxl import Workbook

# 엑셀문서 만들기
wb = Workbook()
print(wb.sheetnames)
ws = wb.active
# 행 순서로 값아 채워진다.
i = 0
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        i += 1
        cell.value = i
        print(cell, end=' ')
    print()
print("~" * 80)
# 열 순서로 값이 채워진다.
i = 0
ws1 = wb.copy_worksheet(ws)
for col in ws1.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        i += 1
        cell.value = i
        print(cell, end=' ')
    print()
print("~" * 80)

# 저장하기
wb.save("Ex06.xlsx")
