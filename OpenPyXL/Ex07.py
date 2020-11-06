# Excel 파일 다루기
from openpyxl import Workbook
import datetime
# 엑셀문서 만들기
wb = Workbook()
ws = wb.active
ws['C9'] = datetime.datetime.today()
# 모든 내용 : A1:C9
print(ws.rows)
print("~" * 80)

for row in ws.rows:
    for cell in row:
        print(cell, end=' ')
    print()
print("~" * 80)
# 값, 형식
print(ws['C9'])
print(ws['C9'].value)
print(ws['C9'].number_format)
print("~" * 80)
# 저장하기
wb.save("Ex07.xlsx")
