# Excel 파일 다루기
from openpyxl import Workbook
import random
# 엑셀문서 만들기
wb = Workbook()
# 워크시트 선택
ws = wb.active
# 값 추가
for row in range(1, 11):
    ws.append(random.sample(list(range(1, 45)), 6))
# 원하는 범위의 값만 읽기 : ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for value in row:
        print(value, end=' ')
    print()
print("~" * 80)
# 저장하기
wb.save("Ex09.xlsx")
